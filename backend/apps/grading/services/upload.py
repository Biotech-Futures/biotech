"""Bulk-mark upload parser + committer.

Accepts an XLSX or CSV that the admin has filled in off-platform and turns it
into a diff against existing ``Grade`` rows for a given component. The
canonical column set is intentionally narrow so the parser stays stable when
the rubric grows or shrinks:

    group_id | criterion_id | mark | comment
             |              |      | (extra columns are ignored, so the
             |              |      |  human-friendly download can carry
             |              |      |  group_name / criterion_name for
             |              |      |  context without breaking parse)

The client is expected to eventually send their own template — this format is
the sensible default from the plan. When their template arrives, only the
column mapping in :func:`_iter_rows` needs to change.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Iterable

from django.db import transaction
from openpyxl import load_workbook

from apps.submissions.models import Submission

from ..models import Grade, RubricCriterion


REQUIRED_COLUMNS = ("group_id", "criterion_id", "mark", "comment")


@dataclass
class UploadDiff:
    creates: list[dict] = field(default_factory=list)
    updates: list[dict] = field(default_factory=list)
    unchanged: list[dict] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)

    def summary(self) -> dict:
        return {
            "creates": len(self.creates),
            "updates": len(self.updates),
            "unchanged": len(self.unchanged),
            "errors": len(self.errors),
        }

    def as_dict(self) -> dict:
        return {
            "creates": self.creates,
            "updates": self.updates,
            "unchanged": self.unchanged,
            "errors": self.errors,
            "summary": self.summary(),
        }


def _iter_rows(file, filename: str) -> Iterable[tuple[int, dict]]:
    """Yield ``(row_number, row_dict)`` for XLSX or CSV inputs.

    Row numbers are 1-indexed and include the header, matching what admins
    see in Excel — makes error messages actionable ("row 12 has a bad mark").
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for i, row in enumerate(reader, start=2):
            yield i, {k.strip(): (v or "").strip() for k, v in row.items() if k}
        return

    # Default: XLSX (openpyxl). Handle both first-row-header and normalised keys.
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return
    keys = [str(h).strip() if h is not None else "" for h in header]
    for i, values in enumerate(rows_iter, start=2):
        row = {}
        for k, v in zip(keys, values):
            if not k:
                continue
            if v is None:
                row[k] = ""
            else:
                row[k] = str(v).strip() if not isinstance(v, str) else v.strip()
        # Skip fully-empty rows (Excel often pads with blank rows after data).
        if not any(row.values()):
            continue
        yield i, row


def _parse_int(value: str) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _parse_mark(value: str) -> tuple[Decimal | None, str | None]:
    """Return (decimal_mark_or_null, error_message_or_null).

    Blank/empty value is a legitimate "clear the mark" and returns (None, None).
    """
    s = str(value or "").strip()
    if s == "":
        return None, None
    try:
        return Decimal(s), None
    except InvalidOperation:
        return None, f"mark {s!r} is not a valid number"


def parse_marks_upload(file, filename: str, component_code: str) -> UploadDiff:
    """Materialise a diff for the given upload against a component's rubric.

    Pure read — no database writes. Callers use ``dry_run`` semantics: show
    the diff to the admin, get confirmation, then re-parse + commit in one
    transaction via :func:`commit_marks_upload`.
    """
    diff = UploadDiff()

    # Preload all valid criterion IDs for this component (rubric.year is not
    # constrained here — bulk uploads could target an older year to correct
    # historical grades; validating structure but not year keeps that path open).
    valid_criteria = {
        c.id: c for c in RubricCriterion.objects.filter(
            rubric__component__code=component_code,
        ).select_related("rubric")
    }
    if not valid_criteria:
        diff.errors.append({"row": 0, "message": f"no rubric criteria exist for component {component_code}"})
        return diff

    submissions_by_group = {
        s.group_id: s for s in Submission.objects.filter(
            component__code=component_code,
        )
    }
    grades_by_pair: dict[tuple[int, int], Grade] = {
        (g.submission.group_id, g.criterion_id): g
        for g in Grade.objects.filter(
            criterion__rubric__component__code=component_code,
        ).select_related("submission", "criterion")
    }

    seen_pairs: set[tuple[int, int]] = set()

    for row_num, row in _iter_rows(file, filename):
        missing = [c for c in REQUIRED_COLUMNS if c not in row]
        if missing:
            diff.errors.append({"row": row_num, "message": f"missing columns: {', '.join(missing)}"})
            continue

        group_id = _parse_int(row["group_id"])
        criterion_id = _parse_int(row["criterion_id"])
        if group_id is None or criterion_id is None:
            diff.errors.append({"row": row_num, "message": "group_id and criterion_id must be integers"})
            continue

        pair = (group_id, criterion_id)
        if pair in seen_pairs:
            diff.errors.append({"row": row_num, "message": f"duplicate (group_id={group_id}, criterion_id={criterion_id})"})
            continue
        seen_pairs.add(pair)

        criterion = valid_criteria.get(criterion_id)
        if criterion is None:
            diff.errors.append({"row": row_num, "message": f"criterion {criterion_id} does not belong to component {component_code}"})
            continue

        submission = submissions_by_group.get(group_id)
        if submission is None:
            diff.errors.append({"row": row_num, "message": f"group {group_id} has no {component_code} submission to grade"})
            continue

        mark, err = _parse_mark(row.get("mark", ""))
        if err:
            diff.errors.append({"row": row_num, "message": err})
            continue
        if mark is not None and mark > criterion.max_mark:
            diff.errors.append({"row": row_num, "message": f"mark {mark} exceeds max_mark {criterion.max_mark}"})
            continue
        if mark is not None and mark < Decimal("0"):
            diff.errors.append({"row": row_num, "message": f"mark {mark} is negative"})
            continue

        comment = row.get("comment", "") or ""
        existing = grades_by_pair.get(pair)
        entry = {
            "row": row_num,
            "group_id": group_id,
            "criterion_id": criterion_id,
            "submission_id": submission.id,
            "mark": str(mark) if mark is not None else None,
            "comment": comment,
        }
        if existing is None:
            diff.creates.append(entry)
        elif existing.mark == mark and (existing.comment or "") == comment:
            diff.unchanged.append({**entry, "grade_id": existing.id})
        else:
            diff.updates.append({
                **entry,
                "grade_id": existing.id,
                "old_mark": str(existing.mark) if existing.mark is not None else None,
                "old_comment": existing.comment or "",
            })

    return diff


@transaction.atomic
def commit_marks_upload(diff: UploadDiff, *, user) -> dict:
    """Apply the ``creates`` and ``updates`` from a validated diff.

    Callers must pre-check ``diff.errors`` is empty — the endpoint refuses to
    commit anything if any row failed validation, so partial imports can't
    happen. ``unchanged`` rows are skipped (no-op, no re-stamping ``graded_by``
    just to record the same value).
    """
    written = 0
    for entry in list(diff.creates) + list(diff.updates):
        Grade.objects.update_or_create(
            submission_id=entry["submission_id"],
            criterion_id=entry["criterion_id"],
            defaults={
                "mark": Decimal(entry["mark"]) if entry["mark"] is not None else None,
                "comment": entry["comment"] or "",
                "graded_by": user,
            },
        )
        written += 1
    return {"written": written, **diff.summary()}
