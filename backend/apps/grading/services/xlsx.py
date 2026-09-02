"""SAQ → XLSX export.

Client explicitly said spreadsheet is easier than PDF for SAQ marking off-
platform, so this is the primary text-export path. Shape:

    | group_id | group_name | submitted_date | submitted_time | is_late | text
    | criterion_1_id | criterion_1_mark | criterion_1_comment
    | criterion_2_id | criterion_2_mark | criterion_2_comment | ...

One row per group; SAQ text goes in a single wrapped cell. Per-criterion
triplets (id + existing mark + existing comment) are appended so the sheet
doubles as a fillable marking template.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ..models import Grade, RubricCriterion
from .content import ComponentEntry


BASE_HEADERS = ["group_id", "group_name", "submitted_date", "submitted_time", "is_late", "text"]


def build_saq_xlsx(
    entries: Iterable[ComponentEntry],
    criteria: Iterable[RubricCriterion] = (),
    grades_by_pair: dict[tuple[int, int], Grade] | None = None,
) -> bytes:
    """Return XLSX bytes for the given SAQ component entries.

    ``criteria`` is the ordered list of rubric criteria for the component;
    each becomes a (criterion_N_id, criterion_N_mark, criterion_N_comment)
    triplet appended to the row. ``grades_by_pair`` maps ``(submission_id,
    criterion_id) -> Grade`` for pre-filling existing marks/comments. Both
    are pushed from the caller so the export layer stays ORM-free.
    """
    criteria_list = list(criteria)
    grades_by_pair = grades_by_pair or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "SAQ"

    headers = list(BASE_HEADERS)
    for i, _ in enumerate(criteria_list, start=1):
        headers.extend([f"criterion_{i}_id", f"criterion_{i}_mark", f"criterion_{i}_comment"])

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        submitted_at = entry.submitted_at
        submitted_date = submitted_at.date().isoformat() if submitted_at else ""
        submitted_time = submitted_at.time().isoformat(timespec="seconds") if submitted_at else ""
        row = [
            entry.group_id,
            entry.group_name,
            submitted_date,
            submitted_time,
            "yes" if entry.is_late else "",
            entry.text or "",
        ]
        for criterion in criteria_list:
            existing = grades_by_pair.get((entry.submission_id, criterion.id))
            row.extend([
                criterion.id,
                float(existing.mark) if existing and existing.mark is not None else None,
                existing.comment if existing else "",
            ])
        ws.append(row)

    # Wrap the text column so long SAQ answers don't just spill off-screen.
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["F"].width = 80

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
