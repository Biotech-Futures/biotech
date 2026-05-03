"""
Load *P11 Test User Data.xlsx* into the database.

  python manage.py import_test_data
  python manage.py import_test_data --dry-run
  python manage.py import_test_data --xlsx /path/to/file.xlsx

Requires a migrated database (``python manage.py migrate``).

Default Excel path: ``<repo-root>/P11 Test User Data.xlsx`` (one level above ``backend/``).
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from apps.groups.models import Countries, CountryStates, Groups, GroupMembership, Tracks
from apps.resources.models import RoleAssignmentHistory, Roles
from apps.users.models import (
    AdminProfile,
    AreasOfInterest,
    MentorProfile,
    StudentProfile,
    SupervisorProfile,
    UserInterest,
)

User = get_user_model()

REGION_TO_TRACK = {
    ("Australia", "NSW"): "AUS-NSW",
    ("Autralia", "NSW"): "AUS-NSW",
    ("Australia", "QLD"): "AUS-QLD",
    ("Australia", "VIC"): "AUS-VIC",
    ("Australia", "WA"): "AUS-WA",
    ("Australia", "SA"): "AUS-SA",
    ("Brazil", None): "BRA",
}
DEFAULT_TRACK = "INTL"

BACKGROUND_MAP = {
    "university student - undergraduate": MentorProfile.BackgroundChoices.UG,
    "university student - postgraduate": MentorProfile.BackgroundChoices.PG,
    "university student - hdr": MentorProfile.BackgroundChoices.HDR,
    "industry": MentorProfile.BackgroundChoices.INDUSTRY,
}


class Command(BaseCommand):
    help = "Import P11 Test User Data.xlsx (synthetic users, groups, roles)"

    def add_arguments(self, parser):
        default_xlsx = Path(settings.BASE_DIR).parent / "P11 Test User Data.xlsx"
        parser.add_argument(
            "--xlsx",
            type=Path,
            default=default_xlsx,
            help=f"Excel file (default: {default_xlsx})",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse the workbook and print counts without writing to the DB",
        )

    def handle(self, *args, **options):
        xlsx: Path = options["xlsx"].resolve()
        dry_run: bool = options["dry_run"]

        if not xlsx.is_file():
            raise CommandError(f"File not found: {xlsx}")

        self.stdout.write(f"Loading {xlsx} …")
        wb = openpyxl.load_workbook(xlsx)

        if dry_run:
            self._dry_run_report(wb)
            return

        self._require_db_tables()

        with transaction.atomic():
            roles = self._ensure_roles()
            track_cache: dict[tuple[str | None, str | None], Tracks] = {}
            group_cache: dict[str, Groups] = {}
            counts = {"supervisors": 0, "admins": 0, "students": 0, "mentors": 0}

            self._import_supervisors(wb["Supervisors"], roles, track_cache, counts)
            self._import_admins(wb["Administrators"], roles, counts)
            self._import_students(wb["Students"], roles, track_cache, group_cache, counts)
            self._import_mentors(wb["Mentors"], roles, track_cache, counts)

        self.stdout.write(
            self.style.SUCCESS(
                f"Import finished — new rows created this run (approx): "
                f"supervisors={counts['supervisors']}, admins={counts['admins']}, "
                f"students={counts['students']}, mentors={counts['mentors']}. "
                f"(Uses get_or_create; re-runs mostly skip existing users.)"
            )
        )

    def _require_db_tables(self) -> None:
        required = ("areas_of_interest", "user_interest", "student_profile", "supervisor_profile")
        missing = []
        with connection.cursor() as cursor:
            for table in required:
                cursor.execute(
                    """
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables
                        WHERE table_schema = 'public' AND table_name = %s
                    )
                    """,
                    [table],
                )
                if not cursor.fetchone()[0]:
                    missing.append(table)
        if missing:
            raise CommandError(
                "Database schema is incomplete — missing table(s): "
                f"{', '.join(missing)}. Run `python manage.py migrate` first."
            )

    def _dry_run_report(self, wb) -> None:
        for name in ("Students", "Mentors", "Supervisors", "Administrators"):
            ws = wb[name]
            n = ws.max_row - 1
            self.stdout.write(f"  {name}: {max(0, n)} data rows")

    def _ensure_roles(self) -> dict[str, Roles]:
        names = ["student", "mentor", "supervisor", "admin"]
        return {n: Roles.objects.get_or_create(role_name=n)[0] for n in names}

    def _get_or_create_track(self, country: str | None, region: str | None, cache):
        key = (country, region)
        if key in cache:
            return cache[key]

        track_name = REGION_TO_TRACK.get(key) or REGION_TO_TRACK.get((country, None))
        if not track_name:
            track_name = DEFAULT_TRACK

        country_name = country or "International"
        state_name = region or track_name

        country_obj, _ = Countries.objects.get_or_create(country_name=country_name)
        state_obj, _ = CountryStates.objects.get_or_create(
            country=country_obj, state_name=state_name
        )
        track_obj, _ = Tracks.objects.get_or_create(
            track_name=track_name,
            defaults={"state": state_obj},
        )
        cache[key] = track_obj
        return track_obj

    def _ensure_track_named(self, track_name: str) -> Tracks:
        """Create a track whose name matches the spreadsheet (e.g. AUS-NSW, BRA)."""
        tname = track_name.strip()
        existing = Tracks.objects.filter(track_name=tname).first()
        if existing:
            return existing

        if tname.startswith("AUS-"):
            state_suffix = tname[4:]
            country_obj, _ = Countries.objects.get_or_create(country_name="Australia")
            state_obj, _ = CountryStates.objects.get_or_create(
                country=country_obj, state_name=state_suffix
            )
        elif tname == "BRA":
            country_obj, _ = Countries.objects.get_or_create(country_name="Brazil")
            state_obj, _ = CountryStates.objects.get_or_create(
                country=country_obj, state_name="Brazil"
            )
        else:
            country_obj, _ = Countries.objects.get_or_create(country_name="International")
            state_obj, _ = CountryStates.objects.get_or_create(
                country=country_obj, state_name=tname
            )

        track_obj, _ = Tracks.objects.get_or_create(
            track_name=tname,
            defaults={"state": state_obj},
        )
        return track_obj

    def _create_or_get_user(self, email, first_name, last_name, track):
        email = (email or "").strip().lower()
        user, created = User.objects.get_or_create(
            email=email,
            defaults={
                "first_name": first_name.strip() if first_name else "",
                "last_name": last_name.strip() if last_name else "",
                "track": track,
                "account_status": User.AccountStatus.INVITED,
                "is_active": False,
            },
        )
        return user, created

    def _assign_role(self, user, role: Roles) -> None:
        RoleAssignmentHistory.objects.get_or_create(
            user=user,
            role=role,
            valid_from=datetime(2025, 1, 1, tzinfo=timezone.utc),
        )

    def _get_or_create_interest(self, name: str) -> AreasOfInterest:
        name = name.strip()
        obj, _ = AreasOfInterest.objects.get_or_create(interest_desc=name)
        return obj

    def _assign_interests(self, user, raw: str | None) -> None:
        if not raw:
            return
        for part in str(raw).split(","):
            part = part.strip()
            if not part:
                continue
            interest = self._get_or_create_interest(part)
            UserInterest.objects.get_or_create(user=user, interest=interest)

    # --- sheets ---

    def _import_supervisors(self, ws, roles, track_cache, counts) -> None:
        default_track = self._get_or_create_track(None, None, track_cache)
        for row in ws.iter_rows(min_row=2, values_only=True):
            first, last, email, school, _conn = row[:5]
            if not email:
                continue
            user, created = self._create_or_get_user(email, first, last, default_track)
            if created:
                counts["supervisors"] += 1
            SupervisorProfile.objects.get_or_create(
                user=user,
                defaults={"school_name": (school or "").strip()},
            )
            self._assign_role(user, roles["supervisor"])

    def _import_admins(self, ws, roles, counts) -> None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            first, last, email, tracks_str = row[:4]
            if not email:
                continue
            track_names = [t.strip() for t in (tracks_str or "").split(",") if t.strip()]
            primary_track = None
            for tname in track_names:
                t = self._ensure_track_named(tname)
                if primary_track is None:
                    primary_track = t
            if primary_track is None:
                primary_track = self._ensure_track_named(DEFAULT_TRACK)
            user, created = self._create_or_get_user(email, first, last, primary_track)
            if created:
                counts["admins"] += 1
            AdminProfile.objects.get_or_create(admin=user)
            self._assign_role(user, roles["admin"])

    def _import_students(self, ws, roles, track_cache, group_cache, counts) -> None:
        supervisor_cache: dict[str, SupervisorProfile | None] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            (
                first,
                last,
                email,
                pg_first,
                pg_last,
                pg_email,
                school,
                year_lvl,
                _interests_old,
                interests,
                group_number,
                _sf,
                _sl,
                sup_email,
                join_perm_response_id,
                country,
                region,
            ) = row[:17]
            if not email:
                continue

            track = self._get_or_create_track(country, region, track_cache)
            user, created = self._create_or_get_user(email, first, last, track)
            if created:
                counts["students"] += 1

            sup_profile = None
            if sup_email:
                key = sup_email.strip().lower()
                if key not in supervisor_cache:
                    try:
                        su = User.objects.get(email=key)
                        sup_profile = SupervisorProfile.objects.filter(user=su).first()
                    except User.DoesNotExist:
                        sup_profile = None
                    supervisor_cache[key] = sup_profile
                sup_profile = supervisor_cache[key]

            year_str = str(int(year_lvl)) if year_lvl is not None else "10"
            if year_str not in {"9", "10", "11", "12"}:
                year_str = "10"

            has_join = bool(join_perm_response_id)
            pg_flag = bool(pg_email) or has_join

            StudentProfile.objects.get_or_create(
                user=user,
                defaults={
                    "pg_first_name": (pg_first or "").strip() or "Unknown",
                    "pg_last_name": (pg_last or "").strip() or "Unknown",
                    "pg_email": (pg_email or "").strip() or None,
                    "parent_guardian_flag": pg_flag,
                    "supervisor": sup_profile,
                    "school_name": (school or "").strip() or "Unknown",
                    "year_lvl": year_str,
                    "has_join_permission": has_join,
                    "joinperm_responseID": join_perm_response_id,
                },
            )
            self._assign_role(user, roles["student"])
            self._assign_interests(user, interests)

            if group_number:
                gid = str(group_number).strip()
                if gid not in group_cache:
                    g, _ = Groups.objects.get_or_create(
                        group_name=gid,
                        defaults={"track": track},
                    )
                    group_cache[gid] = g
                group = group_cache[gid]
                GroupMembership.objects.get_or_create(
                    group=group,
                    user=user,
                    defaults={
                        "membership_role": GroupMembership.MembershipRoleChoices.STUDENT
                    },
                )

    def _import_mentors(self, ws, roles, track_cache, counts) -> None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            (
                first,
                last,
                email,
                interests,
                background_raw,
                institution,
                country,
                region,
                _wwcc,
                _year,
                mentor_reason,
                max_groups,
            ) = row[:12]
            if not email:
                continue

            track = self._get_or_create_track(country, region, track_cache)
            user, created = self._create_or_get_user(email, first, last, track)
            if created:
                counts["mentors"] += 1

            background = BACKGROUND_MAP.get(
                (background_raw or "").strip().lower(),
                MentorProfile.BackgroundChoices.UG,
            )
            reason = (mentor_reason or "").strip()[:255] or "—"

            MentorProfile.objects.get_or_create(
                user=user,
                defaults={
                    "background": background,
                    "institution": (institution or "").strip(),
                    "mentor_reason": reason,
                    "max_group_count": int(max_groups) if max_groups else 2,
                },
            )
            self._assign_role(user, roles["mentor"])
            self._assign_interests(user, interests)
